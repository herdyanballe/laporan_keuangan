"""
Aplikasi Laporan Keuangan Kas Kelompok Narogong
Fitur: Input transaksi, import Excel, export Excel & PDF dengan logo
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import json, os, shutil
from datetime import datetime, date
from pathlib import Path

# ── Data storage ──────────────────────────────────────────────────────────────
DATA_FILE = "data_kas.json"
CONFIG_FILE = "config.json"

BULAN_ID = ["Januari","Februari","Maret","April","Mei","Juni",
            "Juli","Agustus","September","Oktober","November","Desember"]

def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return []

def save_data(data):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"logo_path": "", "nama_kelompok": "Kelompok Narogong", "nama_gereja": "GKJ"}

def save_config(cfg):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

def fmt_rp(val):
    try:
        return f"Rp {int(float(val)):,}".replace(",",".")
    except:
        return "Rp 0"

def parse_tgl(tgl_str):
    """Parse DD-MM-YYYY string, return datetime object or None."""
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(tgl_str)[:10], fmt)
        except:
            continue
    return None

def fmt_tgl(tgl_str):
    """Normalize any date string to DD-MM-YYYY display format."""
    d = parse_tgl(tgl_str)
    return d.strftime("%d-%m-%Y") if d else tgl_str

def tgl_to_sort_key(tgl_str):
    """Return a sort key (datetime) from a date string."""
    d = parse_tgl(tgl_str)
    return d if d else datetime(1900,1,1)

def get_saldo_hari_ini(data):
    """Hitung saldo hingga hari ini"""
    today = date.today()
    saldo = 0
    for r in sorted(data, key=lambda x: tgl_to_sort_key(x.get("tanggal", ""))):
        tgl = parse_tgl(r.get("tanggal", ""))
        if tgl and tgl.date() <= today:
            saldo += r.get("masuk", 0) - r.get("keluar", 0)
    return saldo

# ── Excel Import ───────────────────────────────────────────────────────────────
def import_from_excel(filepath):
    import pandas as pd
    from datetime import datetime as dt
    rows = []
    try:
        df = pd.read_excel(filepath, header=None)
        header_row = None
        for i, row in df.iterrows():
            if "Keterangan" in str(list(row.values)):
                header_row = i
                break
        if header_row is None:
            return [], "Header tidak ditemukan di file Excel."
        df.columns = df.iloc[header_row]
        df = df.iloc[header_row+1:].reset_index(drop=True)
        cols = list(df.columns)
        tgl_col  = next((c for c in cols if "Tanggal" in str(c)), None)
        ket_col  = next((c for c in cols if "Keterangan" in str(c)), None)
        msk_col  = next((c for c in cols if "Masuk" in str(c)), None)
        klr_col  = next((c for c in cols if "Keluar" in str(c)), None)
        if not all([tgl_col, ket_col, msk_col, klr_col]):
            return [], "Kolom tidak lengkap."
        for _, row in df.iterrows():
            ket = str(row[ket_col]).strip()
            if ket in ["nan","","Grand Total","Total","Keterangan"]:
                continue
            tgl = row[tgl_col]
            try:
                if isinstance(tgl, (int, float)):
                    tgl_obj = dt(1899,12,30) + __import__("pandas").Timedelta(days=int(tgl))
                    tgl_str = tgl_obj.strftime("%d-%m-%Y")
                elif hasattr(tgl, "strftime"):
                    tgl_str = tgl.strftime("%d-%m-%Y")
                else:
                    tgl_str = fmt_tgl(str(tgl)[:10])
            except:
                tgl_str = ""
            masuk  = float(row[msk_col]) if str(row[msk_col]) not in ["nan",""] else 0
            keluar = float(row[klr_col]) if str(row[klr_col]) not in ["nan",""] else 0
            rows.append({"tanggal": tgl_str, "keterangan": ket,
                         "masuk": masuk, "keluar": keluar})
    except Exception as e:
        return [], str(e)
    return rows, None

# ── Excel Export ───────────────────────────────────────────────────────────────
def export_excel(data, filepath, config, bulan_filter=None, tahun_filter=None):
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  numbers)
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Kas"

    # Filter data
    rows = filter_data(data, bulan_filter, tahun_filter)

    # ── Header logo area ──
    start_row = 1
    if config.get("logo_path") and os.path.exists(config["logo_path"]):
        try:
            img = XLImage(config["logo_path"])
            img.width, img.height = 70, 70
            img.anchor = "A1"
            ws.add_image(img)
            ws.row_dimensions[1].height = 55
            start_row = 1
        except:
            pass

    # ── Title ──
    title_row = 3
    nama_kel = config.get("nama_kelompok","Kelompok Narogong")
    nama_ger = config.get("nama_gereja","GKJ")
    judul_periode = build_periode_label(rows, bulan_filter, tahun_filter)

    ws.merge_cells(f"B{title_row}:F{title_row}")
    c = ws[f"B{title_row}"]
    c.value = f"LAPORAN KAS {nama_kel.upper()}"
    c.font = Font(name="Arial", bold=True, size=14)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(f"B{title_row+1}:F{title_row+1}")
    c2 = ws[f"B{title_row+1}"]
    c2.value = judul_periode
    c2.font = Font(name="Arial", size=11)
    c2.alignment = Alignment(horizontal="center")

    # ── Column headers ──
    hdr_row = title_row + 3
    headers = ["No","Tanggal","Keterangan","Kas Masuk","Kas Keluar","Saldo Akhir"]
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_widths = [5, 14, 45, 18, 18, 18]
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=hdr_row, column=i, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[hdr_row].height = 28

    # ── Data rows ──
    # Untuk laporan bulanan, saldo dihitung dari saldo awal (sebelum periode) + transaksi dalam periode
    if bulan_filter or tahun_filter:
        # Dapatkan semua data sebelum periode (untuk saldo awal)
        all_data_before = get_data_before_period(data, bulan_filter, tahun_filter)
        saldo_awal = sum(r["masuk"] - r["keluar"] for r in all_data_before)
        saldo = saldo_awal
    else:
        saldo = 0
    
    alt_fill = PatternFill("solid", fgColor="EEF4FB")
    rp_fmt = '#,##0'
    for idx, row in enumerate(rows, 1):
        r = hdr_row + idx
        saldo += row["masuk"] - row["keluar"]
        vals = [idx, row["tanggal"], row["keterangan"],
                row["masuk"] if row["masuk"] else None,
                row["keluar"] if row["keluar"] else None,
                saldo]
        fill = alt_fill if idx % 2 == 0 else None
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            cell.alignment = Alignment(vertical="center",
                                       horizontal="center" if ci in [1,2] else
                                       "left" if ci==3 else "right")
            if fill:
                cell.fill = fill
            if ci in [4,5,6] and v is not None:
                cell.number_format = rp_fmt
        ws.row_dimensions[r].height = 18

    # ── Grand Total ──
    tot_row = hdr_row + len(rows) + 1
    total_masuk  = sum(r["masuk"]  for r in rows)
    total_keluar = sum(r["keluar"] for r in rows)
    
    # Tampilkan saldo awal jika ada filter
    if bulan_filter or tahun_filter:
        all_data_before = get_data_before_period(data, bulan_filter, tahun_filter)
        saldo_awal = sum(r["masuk"] - r["keluar"] for r in all_data_before)
        ws.merge_cells(f"A{tot_row}:C{tot_row}")
        tot_label = ws[f"A{tot_row}"]
        tot_label.value = f"SALDO AWAL: {fmt_rp(saldo_awal)}"
        tot_label.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        tot_label.fill = PatternFill("solid", fgColor="1F4E79")
        tot_label.alignment = Alignment(horizontal="center", vertical="center")
        tot_label.border = border
        
        ws.merge_cells(f"A{tot_row+1}:C{tot_row+1}")
        tot_label2 = ws[f"A{tot_row+1}"]
        tot_label2.value = "GRAND TOTAL"
        tot_label2.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        tot_label2.fill = PatternFill("solid", fgColor="1F4E79")
        tot_label2.alignment = Alignment(horizontal="center", vertical="center")
        tot_label2.border = border
        
        for ci, val in [(4, total_masuk),(5, total_keluar),(6, saldo)]:
            cell = ws.cell(row=tot_row+1, column=ci, value=val)
            cell.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = rp_fmt
            cell.border = border
        ws.row_dimensions[tot_row+1].height = 22
        ws.row_dimensions[tot_row].height = 22
    else:
        ws.merge_cells(f"A{tot_row}:C{tot_row}")
        tot_label = ws[f"A{tot_row}"]
        tot_label.value = "GRAND TOTAL"
        tot_label.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        tot_label.fill = PatternFill("solid", fgColor="1F4E79")
        tot_label.alignment = Alignment(horizontal="center", vertical="center")
        tot_label.border = border

        for ci, val in [(4, total_masuk),(5, total_keluar),(6, saldo)]:
            cell = ws.cell(row=tot_row, column=ci, value=val)
            cell.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = rp_fmt
            cell.border = border
        ws.row_dimensions[tot_row].height = 22

    # ── Freeze header ──
    ws.freeze_panes = f"A{hdr_row+1}"

    wb.save(filepath)

def get_data_before_period(data, bulan_filter, tahun_filter):
    """Ambil semua data sebelum periode filter"""
    result = []
    for r in data:
        tgl = r.get("tanggal", "")
        d = parse_tgl(tgl)
        if d:
            if tahun_filter and d.year < int(tahun_filter):
                result.append(r)
            elif tahun_filter and d.year == int(tahun_filter) and bulan_filter:
                if d.month < int(bulan_filter):
                    result.append(r)
            elif not tahun_filter and not bulan_filter:
                pass
            elif tahun_filter and not bulan_filter:
                if d.year < int(tahun_filter):
                    result.append(r)
    return result

# ── PDF Export ─────────────────────────────────────────────────────────────────
def export_pdf(data, filepath, config, bulan_filter=None, tahun_filter=None):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph, Spacer, Image as RLImage,
                                     HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

    rows = filter_data(data, bulan_filter, tahun_filter)
    nama_kel = config.get("nama_kelompok","Kelompok Narogong")
    judul_periode = build_periode_label(rows, bulan_filter, tahun_filter)

    doc = SimpleDocTemplate(filepath, pagesize=landscape(A4),
                             leftMargin=1.5*cm, rightMargin=1.5*cm,
                             topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    navy   = colors.HexColor("#1F4E79")
    silver = colors.HexColor("#EEF4FB")
    white  = colors.white

    style_title = ParagraphStyle("title", fontName="Helvetica-Bold",
                                  fontSize=14, textColor=navy, alignment=TA_CENTER,
                                  spaceAfter=4)
    style_sub   = ParagraphStyle("sub", fontName="Helvetica",
                                  fontSize=10, textColor=colors.grey,
                                  alignment=TA_CENTER, spaceAfter=2)
    style_cell  = ParagraphStyle("cell", fontName="Helvetica", fontSize=8,
                                  leading=10)

    story = []

    # ── Header: logo + title ──
    logo_path = config.get("logo_path","")
    if logo_path and os.path.exists(logo_path):
        from reportlab.platypus import KeepInFrame
        logo = RLImage(logo_path, width=2*cm, height=2*cm)
        header_data = [[logo,
                        [Paragraph(f"LAPORAN KAS {nama_kel.upper()}", style_title),
                         Paragraph(judul_periode, style_sub)]]]
        header_tbl = Table(header_data, colWidths=[2.5*cm, None])
        header_tbl.setStyle(TableStyle([
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("LEFTPADDING", (0,0), (-1,-1), 0),
        ]))
        story.append(header_tbl)
    else:
        story.append(Paragraph(f"LAPORAN KAS {nama_kel.upper()}", style_title))
        story.append(Paragraph(judul_periode, style_sub))

    story.append(HRFlowable(width="100%", thickness=2, color=navy, spaceAfter=8))

    # ── Table ──
    table_data = [["No","Tanggal","Keterangan","Kas Masuk","Kas Keluar","Saldo Akhir"]]
    
    # Untuk laporan bulanan, saldo dihitung dari saldo awal (sebelum periode) + transaksi dalam periode
    if bulan_filter or tahun_filter:
        all_data_before = get_data_before_period(data, bulan_filter, tahun_filter)
        saldo_awal = sum(r["masuk"] - r["keluar"] for r in all_data_before)
        saldo = saldo_awal
    else:
        saldo = 0
    
    for idx, row in enumerate(rows, 1):
        saldo += row["masuk"] - row["keluar"]
        table_data.append([
            str(idx),
            row["tanggal"],
            Paragraph(row["keterangan"], style_cell),
            fmt_rp(row["masuk"])  if row["masuk"]  else "-",
            fmt_rp(row["keluar"]) if row["keluar"] else "-",
            fmt_rp(saldo),
        ])

    total_masuk  = sum(r["masuk"]  for r in rows)
    total_keluar = sum(r["keluar"] for r in rows)
    
    if bulan_filter or tahun_filter:
        all_data_before = get_data_before_period(data, bulan_filter, tahun_filter)
        saldo_awal = sum(r["masuk"] - r["keluar"] for r in all_data_before)
        table_data.append(["","","", "", "", ""])
        table_data.append(["", "", f"SALDO AWAL: {fmt_rp(saldo_awal)}", "", "", ""])
        table_data.append(["","","GRAND TOTAL", fmt_rp(total_masuk), fmt_rp(total_keluar), fmt_rp(saldo)])
    else:
        table_data.append(["","","GRAND TOTAL", fmt_rp(total_masuk), fmt_rp(total_keluar), fmt_rp(saldo)])

    pw = landscape(A4)[0] - 3*cm
    col_w = [1*cm, 2.2*cm, 10*cm, 3.5*cm, 3.5*cm, 3.5*cm]

    tbl = Table(table_data, colWidths=col_w, repeatRows=1)
    n = len(table_data)
    tbl_style = TableStyle([
        # Header
        ("BACKGROUND",   (0,0), (-1,0), navy),
        ("TEXTCOLOR",    (0,0), (-1,0), white),
        ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,0), 9),
        ("ALIGN",        (0,0), (-1,0), "CENTER"),
        ("VALIGN",       (0,0), (-1,0), "MIDDLE"),
        ("ROWBACKGROUND",(0,0), (-1,0), navy),
        # Data
        ("FONTNAME",     (0,1), (-1,n-2), "Helvetica"),
        ("FONTSIZE",     (0,1), (-1,n-2), 8),
        ("VALIGN",       (0,1), (-1,-1), "MIDDLE"),
        ("ALIGN",        (0,1), (1,n-2), "CENTER"),
        ("ALIGN",        (3,1), (5,n-2), "RIGHT"),
        # Alternating rows
        *[("BACKGROUND", (0,i), (-1,i), silver) for i in range(1, n-1) if i%2==0],
        # Total row
        ("BACKGROUND",   (0,n-1), (-1,n-1), navy),
        ("TEXTCOLOR",    (0,n-1), (-1,n-1), white),
        ("FONTNAME",     (0,n-1), (-1,n-1), "Helvetica-Bold"),
        ("FONTSIZE",     (0,n-1), (-1,n-1), 9),
        ("ALIGN",        (2,n-1), (2,n-1), "CENTER"),
        ("ALIGN",        (3,n-1), (5,n-1), "RIGHT"),
        # Grid
        ("GRID",         (0,0), (-1,-1), 0.4, colors.HexColor("#AAAAAA")),
        ("TOPPADDING",   (0,0), (-1,-1), 4),
        ("BOTTOMPADDING",(0,0), (-1,-1), 4),
        ("LEFTPADDING",  (0,0), (-1,-1), 4),
        ("RIGHTPADDING", (0,0), (-1,-1), 4),
    ])
    tbl.setStyle(tbl_style)
    story.append(tbl)

    # ── Footer ──
    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
    tgl_cetak = datetime.now().strftime("%d %B %Y pukul %H:%M")
    story.append(Paragraph(f"Dicetak pada: {tgl_cetak}", style_sub))

    doc.build(story)

# ── Helpers ────────────────────────────────────────────────────────────────────
def filter_data(data, bulan, tahun):
    result = []
    for r in data:
        tgl = r.get("tanggal","")
        d = parse_tgl(tgl)
        if d:
            if tahun and str(d.year) != str(tahun):
                continue
            if bulan and str(d.month) != str(bulan):
                continue
        result.append(r)
    return sorted(result, key=lambda r: tgl_to_sort_key(r.get("tanggal","")))

def build_periode_label(rows, bulan_filter, tahun_filter):
    if bulan_filter and tahun_filter:
        return f"Bulan {BULAN_ID[int(bulan_filter)-1]} {tahun_filter}"
    if tahun_filter:
        return f"Tahun {tahun_filter}"
    if rows:
        try:
            tanggals = [parse_tgl(r["tanggal"]) for r in rows if r.get("tanggal")]
            tanggals = [d for d in tanggals if d]
            if tanggals:
                mn, mx = min(tanggals), max(tanggals)
                if mn.year == mx.year and mn.month == mx.month:
                    return f"Bulan {BULAN_ID[mn.month-1]} {mn.year}"
                return f"{BULAN_ID[mn.month-1]} {mn.year} s.d. {BULAN_ID[mx.month-1]} {mx.year}"
        except:
            pass
    return "Semua Periode"

# ── Main Application ───────────────────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Laporan Keuangan Kas — Kelompok Narogong")
        self.geometry("1100x750")
        self.configure(bg="#F0F4FA")
        self.resizable(True, True)

        self.data   = load_data()
        self.config = load_config()

        self._build_ui()
        self._refresh_table()

    # ── UI Builder ──────────────────────────────────────────────────────────
    def _build_ui(self):
        # ── Top bar ──
        topbar = tk.Frame(self, bg="#1F4E79", height=56)
        topbar.pack(fill="x")
        topbar.pack_propagate(False)

        tk.Label(topbar, text="💰  Laporan Keuangan Kas",
                 font=("Arial",16,"bold"), bg="#1F4E79", fg="white"
                 ).pack(side="left", padx=18, pady=10)

        tk.Button(topbar, text="⚙  Pengaturan", font=("Arial",10),
                  bg="#2E6DA4", fg="white", relief="flat", cursor="hand2",
                  command=self._open_settings
                  ).pack(side="right", padx=10, pady=12)

        # ── Saldo Hari Ini Panel ──
        saldo_panel = tk.Frame(self, bg="#2E6DA4", height=50)
        saldo_panel.pack(fill="x", padx=12, pady=(8,0))
        saldo_panel.pack_propagate(False)
        
        today = date.today().strftime("%d %B %Y")
        saldo_hari_ini = get_saldo_hari_ini(self.data)
        
        tk.Label(saldo_panel, text=f"📊  SALDO PER TANGGAL {today}  📊",
                 font=("Arial",12,"bold"), bg="#2E6DA4", fg="white"
                 ).pack(side="left", padx=15, pady=12)
        
        self.saldo_hari_ini_label = tk.Label(saldo_panel, text=fmt_rp(saldo_hari_ini),
                                              font=("Arial",16,"bold"), bg="#2E6DA4", 
                                              fg="#FFD700")
        self.saldo_hari_ini_label.pack(side="right", padx=15, pady=12)

        # ── Main area ──
        main = tk.Frame(self, bg="#F0F4FA")
        main.pack(fill="both", expand=True, padx=12, pady=8)

        # Left panel: form
        left = tk.LabelFrame(main, text=" Input Transaksi ", font=("Arial",10,"bold"),
                              bg="#F0F4FA", fg="#1F4E79", padx=10, pady=8)
        left.pack(side="left", fill="y", padx=(0,8))

        fields = [("Tanggal (DD-MM-YYYY):", "tanggal"),
                  ("Keterangan:", "keterangan"),
                  ("Kas Masuk (Rp):", "masuk"),
                  ("Kas Keluar (Rp):", "keluar")]
        self.entries = {}
        for lbl, key in fields:
            tk.Label(left, text=lbl, font=("Arial",9), bg="#F0F4FA",
                     anchor="w").pack(fill="x", pady=(6,0))
            if key == "tanggal":
                var = tk.StringVar(value=date.today().strftime("%d-%m-%Y"))
                e = tk.Entry(left, textvariable=var, font=("Arial",10), width=24)
            elif key == "keterangan":
                e = tk.Text(left, font=("Arial",10), width=24, height=3,
                            relief="solid", bd=1)
                e.bind("<Tab>", lambda ev: self.focus_next_widget(ev))
            else:
                e = tk.Entry(left, font=("Arial",10), width=24)
            if key == "tanggal":
                e.config(relief="solid", bd=1)
            if key != "keterangan":
                e.pack(fill="x")
            else:
                e.pack(fill="x")
            self.entries[key] = e if key != "tanggal" else (e, var)

        tk.Button(left, text="✚  Tambah Transaksi", font=("Arial",10,"bold"),
                  bg="#1F4E79", fg="white", relief="flat", cursor="hand2",
                  pady=8, command=self._add_entry
                  ).pack(fill="x", pady=(14,4))

        tk.Button(left, text="📂  Import Excel", font=("Arial",10),
                  bg="#2E6DA4", fg="white", relief="flat", cursor="hand2",
                  pady=6, command=self._import_excel
                  ).pack(fill="x", pady=2)

        tk.Button(left, text="🗑  Hapus Dipilih", font=("Arial",10),
                  bg="#C0392B", fg="white", relief="flat", cursor="hand2",
                  pady=6, command=self._delete_selected
                  ).pack(fill="x", pady=2)

        # Right panel: table + filter + export
        right = tk.Frame(main, bg="#F0F4FA")
        right.pack(side="left", fill="both", expand=True)

        # Filter row
        flt = tk.Frame(right, bg="#F0F4FA")
        flt.pack(fill="x", pady=(0,6))

        tk.Label(flt, text="Filter Bulan:", font=("Arial",9), bg="#F0F4FA").pack(side="left")
        self.filter_bulan = ttk.Combobox(flt, width=12, font=("Arial",9),
                                          values=["Semua"]+BULAN_ID, state="readonly")
        self.filter_bulan.set("Semua")
        self.filter_bulan.pack(side="left", padx=(4,12))

        tk.Label(flt, text="Tahun:", font=("Arial",9), bg="#F0F4FA").pack(side="left")
        years = sorted({str(parse_tgl(r["tanggal"]).year) for r in self.data
                        if r.get("tanggal") and parse_tgl(r["tanggal"])}, reverse=True) or [str(date.today().year)]
        self.filter_tahun = ttk.Combobox(flt, width=8, font=("Arial",9),
                                          values=["Semua"]+years, state="readonly")
        self.filter_tahun.set("Semua")
        self.filter_tahun.pack(side="left", padx=(4,12))

        tk.Button(flt, text="🔍 Tampilkan", font=("Arial",9),
                  bg="#1F4E79", fg="white", relief="flat", cursor="hand2",
                  command=self._refresh_table).pack(side="left", padx=4)

        # Export buttons
        tk.Button(flt, text="📊 Export Excel", font=("Arial",9,"bold"),
                  bg="#217346", fg="white", relief="flat", cursor="hand2",
                  command=self._export_excel).pack(side="right", padx=4)
        tk.Button(flt, text="📄 Export PDF", font=("Arial",9,"bold"),
                  bg="#C0392B", fg="white", relief="flat", cursor="hand2",
                  command=self._export_pdf).pack(side="right", padx=4)

        # Treeview
        cols = ("no","tanggal","keterangan","masuk","keluar","saldo")
        self.tree = ttk.Treeview(right, columns=cols, show="headings",
                                  selectmode="browse")
        hdrs = [("no","No",40),("tanggal","Tanggal",100),
                ("keterangan","Keterangan",360),
                ("masuk","Kas Masuk",130),("keluar","Kas Keluar",130),
                ("saldo","Saldo Akhir",130)]
        for cid, lbl, w in hdrs:
            self.tree.heading(cid, text=lbl)
            anchor = "e" if cid in ("masuk","keluar","saldo") else \
                     "center" if cid in ("no","tanggal") else "w"
            self.tree.column(cid, width=w, anchor=anchor)

        style = ttk.Style()
        style.configure("Treeview", font=("Arial",9), rowheight=22)
        style.configure("Treeview.Heading", font=("Arial",9,"bold"),
                         background="#1F4E79", foreground="white")

        vsb = ttk.Scrollbar(right, orient="vertical",   command=self.tree.yview)
        hsb = ttk.Scrollbar(right, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

        # Status bar
        self.status_var = tk.StringVar(value="Siap.")
        tk.Label(self, textvariable=self.status_var, font=("Arial",9),
                 bg="#1F4E79", fg="white", anchor="w", padx=10
                 ).pack(fill="x", side="bottom")

    # ── Actions ─────────────────────────────────────────────────────────────
    def focus_next_widget(self, event):
        event.widget.tk_focusNext().focus()
        return "break"

    def _get_filter(self):
        b = self.filter_bulan.get()
        t = self.filter_tahun.get()
        bulan = str(BULAN_ID.index(b)+1) if b != "Semua" else None
        tahun = t if t != "Semua" else None
        return bulan, tahun

    def _refresh_table(self):
        bulan, tahun = self._get_filter()
        rows = filter_data(self.data, bulan, tahun)

        self.tree.delete(*self.tree.get_children())
        
        # Hitung saldo dengan mempertimbangkan filter
        if bulan or tahun:
            # Dapatkan saldo awal dari data sebelum periode
            all_data_before = get_data_before_period(self.data, bulan, tahun)
            saldo = sum(r["masuk"] - r["keluar"] for r in all_data_before)
        else:
            saldo = 0
            
        for idx, row in enumerate(rows, 1):
            saldo += row["masuk"] - row["keluar"]
            tag = "even" if idx % 2 == 0 else "odd"
            self.tree.insert("", "end", iid=str(idx-1),
                              values=(idx, row["tanggal"], row["keterangan"],
                                      fmt_rp(row["masuk"])  if row["masuk"]  else "-",
                                      fmt_rp(row["keluar"]) if row["keluar"] else "-",
                                      fmt_rp(saldo)),
                              tags=(tag,))
        self.tree.tag_configure("even", background="#EEF4FB")
        self.tree.tag_configure("odd",  background="#FFFFFF")

        total_masuk  = sum(r["masuk"]  for r in rows)
        total_keluar = sum(r["keluar"] for r in rows)
        
        # Update status bar
        if bulan or tahun:
            all_data_before = get_data_before_period(self.data, bulan, tahun)
            saldo_awal = sum(r["masuk"] - r["keluar"] for r in all_data_before)
            self.status_var.set(
                f"  {len(rows)} transaksi  |  Saldo Awal: {fmt_rp(saldo_awal)}  "
                f"|  Total Masuk: {fmt_rp(total_masuk)}  "
                f"|  Total Keluar: {fmt_rp(total_keluar)}  |  Saldo Akhir: {fmt_rp(saldo)}"
            )
        else:
            self.status_var.set(
                f"  {len(rows)} transaksi  |  Total Masuk: {fmt_rp(total_masuk)}  "
                f"|  Total Keluar: {fmt_rp(total_keluar)}  |  Saldo Akhir: {fmt_rp(saldo)}"
            )
        
        # Update saldo hari ini
        saldo_hari_ini = get_saldo_hari_ini(self.data)
        self.saldo_hari_ini_label.config(text=fmt_rp(saldo_hari_ini))

    def _add_entry(self):
        tgl_entry, tgl_var = self.entries["tanggal"]
        tgl = tgl_var.get().strip()
        ket = self.entries["keterangan"].get("1.0","end").strip()
        masuk_str  = self.entries["masuk"].get().strip().replace(".","").replace(",","")
        keluar_str = self.entries["keluar"].get().strip().replace(".","").replace(",","")

        if not tgl or not ket:
            messagebox.showwarning("Input Kosong", "Tanggal dan Keterangan wajib diisi.")
            return
        if not parse_tgl(tgl):
            messagebox.showerror("Format Tanggal", "Format tanggal harus DD-MM-YYYY\nContoh: 25-06-2024")
            return
        tgl = fmt_tgl(tgl)  # normalize to DD-MM-YYYY

        masuk  = float(masuk_str)  if masuk_str  else 0
        keluar = float(keluar_str) if keluar_str else 0

        self.data.append({"tanggal":tgl,"keterangan":ket,"masuk":masuk,"keluar":keluar})
        save_data(self.data)
        self._update_year_filter()
        self._refresh_table()
        # Clear form
        self.entries["keterangan"].delete("1.0","end")
        self.entries["masuk"].delete(0,"end")
        self.entries["keluar"].delete(0,"end")
        self.status_var.set("  ✅ Transaksi berhasil ditambahkan.")

    def _delete_selected(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Pilih Dulu", "Pilih baris yang ingin dihapus.")
            return
        idx = int(sel[0])
        bulan, tahun = self._get_filter()
        rows = filter_data(self.data, bulan, tahun)
        row_to_del = rows[idx]
        if messagebox.askyesno("Konfirmasi", f"Hapus transaksi:\n{row_to_del['keterangan']}?"):
            self.data.remove(row_to_del)
            save_data(self.data)
            self._refresh_table()

    def _import_excel(self):
        path = filedialog.askopenfilename(
            title="Pilih file Excel",
            filetypes=[("Excel files","*.xlsx *.xls")]
        )
        if not path:
            return
        rows, err = import_from_excel(path)
        if err:
            messagebox.showerror("Error Import", err)
            return
        existing_keys = {(r["tanggal"],r["keterangan"],r["masuk"],r["keluar"])
                         for r in self.data}
        added = 0
        for r in rows:
            key = (r["tanggal"],r["keterangan"],r["masuk"],r["keluar"])
            if key not in existing_keys:
                self.data.append(r)
                existing_keys.add(key)
                added += 1
        save_data(self.data)
        self._update_year_filter()
        self._refresh_table()
        messagebox.showinfo("Import Selesai",
                             f"Berhasil import {added} transaksi baru.\n"
                             f"({len(rows)-added} transaksi duplikat dilewati)")

    def _update_year_filter(self):
        years = sorted({str(parse_tgl(r["tanggal"]).year) for r in self.data
                        if r.get("tanggal") and parse_tgl(r["tanggal"])}, reverse=True)
        self.filter_tahun["values"] = ["Semua"] + years

    def _export_excel(self):
        bulan, tahun = self._get_filter()
        rows = filter_data(self.data, bulan, tahun)
        if not rows:
            messagebox.showwarning("Data Kosong", "Tidak ada data untuk diekspor.")
            return
        default = f"Laporan_Kas_{build_periode_label(rows,bulan,tahun).replace(' ','_')}.xlsx"
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", initialfile=default,
            filetypes=[("Excel","*.xlsx")]
        )
        if not path:
            return
        try:
            export_excel(self.data, path, self.config, bulan, tahun)
            messagebox.showinfo("Sukses", f"File Excel berhasil disimpan:\n{path}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _export_pdf(self):
        bulan, tahun = self._get_filter()
        rows = filter_data(self.data, bulan, tahun)
        if not rows:
            messagebox.showwarning("Data Kosong", "Tidak ada data untuk diekspor.")
            return
        default = f"Laporan_Kas_{build_periode_label(rows,bulan,tahun).replace(' ','_')}.pdf"
        path = filedialog.asksaveasfilename(
            defaultextension=".pdf", initialfile=default,
            filetypes=[("PDF","*.pdf")]
        )
        if not path:
            return
        try:
            export_pdf(self.data, path, self.config, bulan, tahun)
            messagebox.showinfo("Sukses", f"File PDF berhasil disimpan:\n{path}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _open_settings(self):
        win = tk.Toplevel(self)
        win.title("Pengaturan")
        win.geometry("420x260")
        win.configure(bg="#F0F4FA")
        win.grab_set()

        tk.Label(win, text="Pengaturan Aplikasi", font=("Arial",12,"bold"),
                  bg="#F0F4FA", fg="#1F4E79").pack(pady=(16,10))

        frm = tk.Frame(win, bg="#F0F4FA")
        frm.pack(padx=20, fill="x")

        for row_idx, (label, key) in enumerate([
            ("Nama Kelompok:", "nama_kelompok"),
            ("Nama Gereja:", "nama_gereja"),
        ]):
            tk.Label(frm, text=label, font=("Arial",9), bg="#F0F4FA",
                      width=16, anchor="w").grid(row=row_idx, column=0, pady=6, sticky="w")
            var = tk.StringVar(value=self.config.get(key,""))
            e = tk.Entry(frm, textvariable=var, font=("Arial",10),
                          width=28, relief="solid", bd=1)
            e.grid(row=row_idx, column=1, padx=6, sticky="w")
            frm.__dict__[key] = var

        # Logo picker
        tk.Label(frm, text="Logo (PNG/JPG):", font=("Arial",9), bg="#F0F4FA",
                  width=16, anchor="w").grid(row=2, column=0, pady=6, sticky="w")
        logo_var = tk.StringVar(value=self.config.get("logo_path",""))
        logo_entry = tk.Entry(frm, textvariable=logo_var, font=("Arial",9),
                               width=22, relief="solid", bd=1, state="readonly")
        logo_entry.grid(row=2, column=1, padx=6, sticky="w")
        def pick_logo():
            p = filedialog.askopenfilename(
                title="Pilih Logo", filetypes=[("Image","*.png *.jpg *.jpeg")])
            if p:
                # Copy to app folder
                dest = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                    "logo" + Path(p).suffix)
                shutil.copy2(p, dest)
                logo_var.set(dest)
        tk.Button(frm, text="Pilih...", font=("Arial",9), command=pick_logo,
                   relief="solid", bd=1).grid(row=2, column=2, padx=4)

        def save_settings():
            self.config["nama_kelompok"] = frm.nama_kelompok.get()
            self.config["nama_gereja"]   = frm.nama_gereja.get()
            self.config["logo_path"]     = logo_var.get()
            save_config(self.config)
            win.destroy()
            messagebox.showinfo("Tersimpan", "Pengaturan berhasil disimpan.")

        tk.Button(win, text="💾  Simpan Pengaturan", font=("Arial",10,"bold"),
                   bg="#1F4E79", fg="white", relief="flat", cursor="hand2",
                   pady=8, command=save_settings
                   ).pack(pady=16, padx=20, fill="x")

# ── Entry point ────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = App()
    app.mainloop()